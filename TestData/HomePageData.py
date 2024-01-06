import openpyxl

class HomePageData:
    test_HomePage_Data = [{'FirstName': 'Atul', 'LastName': 'Arya', 'HomeTown': 'Kanpur'},
                          {'FirstName': 'valishal', 'LastName': 'Gupta', 'HomeTown': 'prayagraj'}]

    @staticmethod  # this is just like static keyword in java no need to create object simply call using className
    def getExcelData():
        Dict={}
        book = openpyxl.load_workbook("Financial_Sample.xlsx")
        sheet = book.active
        list = []
        for i in range(1, 3 + 1):

            for j in range(1, 4 + 1):
                # print(sheet.cell(row=i, column=j).value)
                # here we are assuming keys are stored in the first row and values are in further below rows
                dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
                list.append(dict)
        return list # this will return list of dictionaries

