# Financial Sample
import openpyxl

book = openpyxl.load_workbook("Financial_Sample.xlsx")
sheet = book.active
cell = sheet.cell(row=5, column=2)
print(f'>>>>>>>>>>>{cell.value}')
print('hellp')

# writing into the sheet and then reading the value
sheet.cell(row=2, column=2).value = 'atul >>>'
cell = sheet.cell(row=2, column=2)
print({cell.value})

# check the length at the sheet level
print(f' total no. of rows {sheet.max_row}')
print(f'total no. of collms {sheet.max_column}')

# another way to extract the data from cell

print(f' data in B4 >>> {sheet["B4"].value}')

# looping all the data from excel
print('++++++++++++++++++++++++++++')

dict ={}
list=[]
for i in range(1, 3+1):

        for j in range(1, 4+1):
            # print(sheet.cell(row=i, column=j).value)
            # here we are assuming keys are stored in the first row and values are in further below rows
            dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            list.append(dict)

print(list)
